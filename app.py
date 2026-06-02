import csv
import io
import json
import os
import sqlite3
import sys
import threading
import time
import traceback
import uuid
from decimal import Decimal, InvalidOperation, ROUND_CEILING, ROUND_FLOOR, getcontext
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font
from web3 import Web3
from wallet_report import generate_wallet_report

getcontext().prec = 50


def get_resource_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS", Path(sys.executable).resolve().parent))
    return Path(__file__).resolve().parent


def get_data_dir() -> Path:
    override = os.environ.get("TOKEN_SCANNER_DATA_DIR", "").strip()
    if override:
        return Path(override).expanduser().resolve()
    if getattr(sys, "frozen", False):
        local_appdata = os.environ.get("LOCALAPPDATA") or str(Path.home() / "AppData" / "Local")
        return Path(local_appdata) / "TokenScannerLocal"
    return Path(__file__).resolve().parent


RESOURCE_DIR = get_resource_dir()
DATA_DIR = get_data_dir()
DATA_DIR.mkdir(parents=True, exist_ok=True)

TOKENS_PATH = RESOURCE_DIR / "tokens.json"
DB_PATH = DATA_DIR / "scanner_data.db"
WALLET_REPORTS_DIR = DATA_DIR / "wallet_reports"
WALLET_REPORTS_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(
    __name__,
    template_folder=str(RESOURCE_DIR / "templates"),
    static_folder=str(RESOURCE_DIR / "static"),
)

RPC_URL_DEFAULT = "https://rpc1.netsbo.io"
ZERO_ADDR = "0x0000000000000000000000000000000000000000"
PAGE_SIZE = max(int(os.environ.get("TOKEN_SCANNER_PAGE_SIZE", "50")), 1)
SCAN_BATCH_SIZE = max(int(os.environ.get("TOKEN_SCANNER_BATCH_SIZE", "5000")), 1)
SCAN_MIN_BATCH_SIZE = max(int(os.environ.get("TOKEN_SCANNER_MIN_BATCH_SIZE", "100")), 1)
TOKEN_SCANNER_CONFIRMATION_BLOCKS = max(int(os.environ.get("TOKEN_SCANNER_CONFIRMATION_BLOCKS", "6")), 0)
TOKEN_SCANNER_REORG_LOOKBACK_BLOCKS = max(
    int(os.environ.get("TOKEN_SCANNER_REORG_LOOKBACK_BLOCKS", str(max(TOKEN_SCANNER_CONFIRMATION_BLOCKS, 25)))),
    TOKEN_SCANNER_CONFIRMATION_BLOCKS,
)
STRICT_RESULT_VALIDATION_DEFAULT = os.environ.get("TOKEN_SCANNER_STRICT_RESULT_VALIDATION", "").lower() in {
    "1", "true", "yes", "on"
}
JOB_RETENTION_SECONDS = max(int(os.environ.get("TOKEN_SCANNER_JOB_RETENTION_SECONDS", "3600")), 60)
SQLITE_TIMEOUT_SECONDS = max(float(os.environ.get("TOKEN_SCANNER_SQLITE_TIMEOUT", "30")), 1.0)
BALANCE_NUMBER_FORMAT = "0.########################"
DEBUG_MODE = os.environ.get("FLASK_DEBUG", "").lower() in {"1", "true", "yes"}
WALLET_REPORT_BATCH_SIZE = max(int(os.environ.get("TOKEN_SCANNER_WALLET_REPORT_BATCH_SIZE", "3000")), 1)
WALLET_REPORT_MIN_BATCH_SIZE = max(int(os.environ.get("TOKEN_SCANNER_WALLET_REPORT_MIN_BATCH_SIZE", "100")), 1)
WALLET_REPORT_MAX_DAYS = max(int(os.environ.get("TOKEN_SCANNER_WALLET_REPORT_MAX_DAYS", "3650")), 1)

DECIMALS_ABI = [{
    "constant": True, "inputs": [], "name": "decimals",
    "outputs": [{"name": "", "type": "uint8"}],
    "type": "function"
}]

BALANCEOF_ABI = [{
    "constant": True,
    "inputs": [{"name": "_owner", "type": "address"}],
    "name": "balanceOf",
    "outputs": [{"name": "balance", "type": "uint256"}],
    "payable": False,
    "stateMutability": "view",
    "type": "function"
}]

JOBS: dict[str, dict[str, Any]] = {}
JOBS_LOCK = threading.Lock()
TOKEN_CACHE_LOCK = threading.Lock()
TOKEN_CACHE: dict[str, Any] = {"mtime": None, "tokens": [], "by_symbol": {}}
DB_INIT_LOCK = threading.Lock()
DB_INITIALIZED = False


class JobCancelled(Exception):
    pass


def get_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=SQLITE_TIMEOUT_SECONDS, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=FULL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    global DB_INITIALIZED
    with DB_INIT_LOCK:
        if DB_INITIALIZED:
            return
        conn = get_db_connection()
        try:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS token_index_state (
                    token_symbol TEXT PRIMARY KEY,
                    contract_address TEXT NOT NULL,
                    decimals INTEGER NOT NULL,
                    start_block INTEGER NOT NULL,
                    last_scanned_block INTEGER NOT NULL,
                    latest_known_block INTEGER,
                    total_transfers INTEGER NOT NULL DEFAULT 0,
                    total_holders INTEGER NOT NULL DEFAULT 0,
                    last_scanned_at REAL,
                    last_error TEXT
                );

                CREATE TABLE IF NOT EXISTS token_transfer_events (
                    token_symbol TEXT NOT NULL,
                    tx_hash TEXT NOT NULL,
                    log_index INTEGER NOT NULL,
                    block_number INTEGER NOT NULL,
                    block_hash TEXT,
                    from_address TEXT NOT NULL,
                    to_address TEXT NOT NULL,
                    raw_value TEXT NOT NULL,
                    PRIMARY KEY (token_symbol, tx_hash, log_index),
                    FOREIGN KEY (token_symbol) REFERENCES token_index_state(token_symbol) ON DELETE CASCADE
                );

                CREATE INDEX IF NOT EXISTS idx_token_transfer_events_token_block
                ON token_transfer_events(token_symbol, block_number);

                CREATE TABLE IF NOT EXISTS token_holders (
                    token_symbol TEXT NOT NULL,
                    wallet TEXT NOT NULL,
                    raw_balance TEXT NOT NULL,
                    raw_balance_length INTEGER NOT NULL,
                    updated_at REAL NOT NULL,
                    PRIMARY KEY (token_symbol, wallet),
                    FOREIGN KEY (token_symbol) REFERENCES token_index_state(token_symbol) ON DELETE CASCADE
                );

                CREATE INDEX IF NOT EXISTS idx_token_holders_token_balance
                ON token_holders(token_symbol, raw_balance_length DESC, raw_balance DESC);
                """
            )
            conn.commit()
            DB_INITIALIZED = True
        finally:
            conn.close()


def load_tokens() -> list[dict[str, Any]]:
    mtime = TOKENS_PATH.stat().st_mtime
    with TOKEN_CACHE_LOCK:
        if TOKEN_CACHE["mtime"] == mtime:
            return TOKEN_CACHE["tokens"]

        with TOKENS_PATH.open("r", encoding="utf-8") as f:
            raw_tokens = json.load(f)

        if not isinstance(raw_tokens, list):
            raise ValueError("tokens.json debe contener una lista de tokens")

        tokens: list[dict[str, Any]] = []
        by_symbol: dict[str, dict[str, Any]] = {}
        for item in raw_tokens:
            if not isinstance(item, dict):
                continue

            symbol = str(item.get("symbol", "")).strip()
            contract_address = str(item.get("contractAddress", "")).strip()
            if not symbol or not contract_address:
                continue

            token = dict(item)
            token["symbol"] = symbol
            token["name"] = str(item.get("name", symbol)).strip() or symbol
            token["contractAddress"] = contract_address
            token["decimals"] = int(item.get("decimals", 18))
            token["startBlock"] = max(int(item.get("startBlock", 0) or 0), 0)
            tokens.append(token)
            by_symbol[symbol.lower()] = token

        if not tokens:
            raise ValueError("tokens.json no contiene tokens válidos")

        TOKEN_CACHE["mtime"] = mtime
        TOKEN_CACHE["tokens"] = tokens
        TOKEN_CACHE["by_symbol"] = by_symbol
        return tokens


def get_token_by_symbol(symbol: str) -> dict[str, Any] | None:
    normalized_symbol = symbol.strip().lower()
    if not normalized_symbol:
        return None
    load_tokens()
    with TOKEN_CACHE_LOCK:
        return TOKEN_CACHE["by_symbol"].get(normalized_symbol)


def cleanup_finished_jobs(now: float | None = None):
    current_time = now or time.time()
    cutoff = current_time - JOB_RETENTION_SECONDS
    with JOBS_LOCK:
        stale_job_ids = [
            job_id
            for job_id, job in JOBS.items()
            if job.get("status") in {"done", "error", "cancelled"}
            and job.get("finishedAt", job.get("createdAt", current_time)) < cutoff
        ]
        for job_id in stale_job_ids:
            JOBS.pop(job_id, None)


def get_web3_connection() -> Web3:
    rpc_url = os.environ.get("NETSBO_RPC_URL", RPC_URL_DEFAULT)
    web3 = Web3(Web3.HTTPProvider(rpc_url, request_kwargs={"timeout": 120}))
    try:
        from web3.middleware import ExtraDataToPOAMiddleware
        web3.middleware_onion.inject(ExtraDataToPOAMiddleware, layer=0)
    except Exception:
        try:
            from web3.middleware import geth_poa_middleware
            web3.middleware_onion.inject(geth_poa_middleware, layer=0)
        except Exception:
            pass
    if not web3.is_connected():
        raise ConnectionError(f"No se pudo conectar al RPC de Netsbo en {rpc_url}")
    return web3


def update_job(job_id: str, **fields):
    with JOBS_LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(fields)


def get_job_snapshot(job_id: str) -> dict[str, Any] | None:
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        return dict(job) if job else None


def get_wallet_report_manifest_path(job_id: str) -> Path:
    return WALLET_REPORTS_DIR / f"{job_id}.json"


def get_wallet_report_file_path(job_id: str) -> Path:
    return WALLET_REPORTS_DIR / f"{job_id}.xlsx"


def build_wallet_report_download_url(job_id: str) -> str:
    return f"/api/wallet-report/download?job_id={job_id}"


def persist_wallet_report_manifest(job: dict[str, Any]):
    manifest = {
        "jobId": job.get("jobId"),
        "jobType": "wallet-report",
        "status": job.get("status"),
        "phase": job.get("phase"),
        "progress": job.get("progress", 0),
        "message": job.get("message"),
        "error": job.get("error"),
        "wallet": job.get("wallet"),
        "days": job.get("days"),
        "startBlock": job.get("startBlock"),
        "endBlock": job.get("endBlock"),
        "totalBlocks": job.get("totalBlocks"),
        "totalTransfers": job.get("totalTransfers", 0),
        "transfersIn": job.get("transfersIn", 0),
        "transfersOut": job.get("transfersOut", 0),
        "transfersSelf": job.get("transfersSelf", 0),
        "uniqueTokens": job.get("uniqueTokens", 0),
        "reportFilename": job.get("reportFilename"),
        "downloadUrl": job.get("downloadUrl"),
        "createdAt": job.get("createdAt"),
        "startedAt": job.get("startedAt"),
        "finishedAt": job.get("finishedAt"),
        "durationSeconds": job.get("durationSeconds"),
        "cancelRequested": bool(job.get("cancelRequested", False)),
    }
    manifest_path = get_wallet_report_manifest_path(str(job.get("jobId", "")))
    temp_path = manifest_path.with_suffix(".json.tmp")
    temp_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(manifest_path)


def load_wallet_report_manifest(job_id: str) -> dict[str, Any] | None:
    manifest_path = get_wallet_report_manifest_path(job_id)
    if not manifest_path.exists():
        return None
    try:
        return json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return None


def get_wallet_report_job_snapshot(job_id: str) -> dict[str, Any] | None:
    job = get_job_snapshot(job_id)
    if job and job.get("jobType") == "wallet-report":
        return job
    return load_wallet_report_manifest(job_id)


def ensure_job_not_cancelled(job_id: str | None):
    if not job_id:
        return
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            raise JobCancelled("El trabajo ya no está disponible.")
        if job.get("cancelRequested"):
            raise JobCancelled("Escaneo cancelado por el usuario.")


def parse_optional_decimal(value: Any, field_name: str) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} debe ser un número válido") from exc
    if not parsed.is_finite():
        raise ValueError(f"{field_name} debe ser un número finito")
    if parsed < 0:
        raise ValueError(f"{field_name} no puede ser negativo")
    return parsed


def parse_optional_bool(value: Any, field_name: str) -> bool | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "on", "si", "sí"}:
        return True
    if normalized in {"0", "false", "no", "off"}:
        return False
    raise ValueError(f"{field_name} debe ser booleano")


def parse_page_param(value: Any) -> int:
    try:
        page = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("page debe ser un entero positivo") from exc
    if page < 1:
        raise ValueError("page debe ser un entero positivo")
    return page


def decimal_to_plain_string(value: Decimal) -> str:
    text = format(value.normalize(), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def raw_to_decimal(raw_value: str, decimals: int) -> Decimal:
    return Decimal(raw_value) / (Decimal(10) ** decimals)


def serialize_results(wallets: list[tuple[str, Decimal]], token_symbol: str) -> list[dict[str, str]]:
    return [
        {
            "wallet": wallet,
            "balance": decimal_to_plain_string(balance),
            "tokenSymbol": token_symbol,
        }
        for wallet, balance in wallets
    ]


def get_token_decimals(web3: Web3, token_address: str, fallback_decimals: int | None = None) -> int:
    contract = web3.eth.contract(address=token_address, abi=DECIMALS_ABI)
    try:
        return int(contract.functions.decimals().call())
    except Exception:
        return int(fallback_decimals or 18)


def balance_of_with_retry(contract, wallet: str, job_id: str | None, max_retries: int = 4) -> int:
    backoff_times = [0.5, 1.0, 2.0, 4.0]
    attempt = 0
    while True:
        ensure_job_not_cancelled(job_id)
        try:
            return int(contract.functions.balanceOf(wallet).call())
        except Exception:
            attempt += 1
            if attempt <= max_retries:
                time.sleep(backoff_times[min(attempt - 1, len(backoff_times) - 1)])
                continue
            raise


def safe_get_logs(
    web3: Web3,
    token_address: str,
    transfer_topic: str,
    from_block: int,
    to_block: int,
    job_id: str | None,
    max_retries: int = 6,
    min_batch: int = 100,
) -> list:
    assert from_block <= to_block
    attempt = 0
    while True:
        ensure_job_not_cancelled(job_id)
        try:
            return web3.eth.get_logs({
                "fromBlock": from_block,
                "toBlock": to_block,
                "address": token_address,
                "topics": [transfer_topic],
            })
        except Exception as exc:
            attempt += 1
            span = to_block - from_block + 1
            if span > min_batch:
                mid = from_block + (span // 2) - 1
                left = safe_get_logs(
                    web3, token_address, transfer_topic, from_block, mid, job_id,
                    max_retries=max_retries, min_batch=min_batch,
                )
                right = safe_get_logs(
                    web3, token_address, transfer_topic, mid + 1, to_block, job_id,
                    max_retries=max_retries, min_batch=min_batch,
                )
                return left + right
            if attempt <= max_retries:
                time.sleep(min(10, 0.8 * attempt + (attempt ** 1.2) * 0.15))
                continue
            raise RuntimeError(
                f"Fallo persistente de get_logs en el rango {from_block}-{to_block}. Error: {exc}"
            ) from exc


def to_hex_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        return Web3.to_hex(value)
    except Exception:
        if hasattr(value, "hex"):
            return "0x" + value.hex()
        return str(value)


def to_int_from_hex(value: Any) -> int:
    hex_value = to_hex_string(value)
    if hex_value.startswith("0x"):
        hex_value = hex_value[2:]
    return int(hex_value or "0", 16)


def decode_transfer_log(log: Any) -> dict[str, Any] | None:
    topics = log.get("topics", [])
    if len(topics) < 3:
        return None
    from_addr = "0x" + topics[1].hex()[-40:]
    to_addr = "0x" + topics[2].hex()[-40:]
    try:
        from_addr = Web3.to_checksum_address(from_addr)
        to_addr = Web3.to_checksum_address(to_addr)
    except Exception:
        return None

    tx_hash = to_hex_string(log.get("transactionHash"))
    block_hash = to_hex_string(log.get("blockHash"))
    raw_value = to_int_from_hex(log.get("data"))
    log_index = int(log.get("logIndex", 0))
    block_number = int(log.get("blockNumber", 0))
    return {
        "tx_hash": tx_hash,
        "block_hash": block_hash,
        "log_index": log_index,
        "block_number": block_number,
        "from_address": from_addr,
        "to_address": to_addr,
        "raw_value": raw_value,
    }


def reset_token_index(
    conn: sqlite3.Connection,
    token_symbol: str,
    contract_address: str,
    decimals: int,
    start_block: int,
    reason: str | None = None,
):
    conn.execute("DELETE FROM token_transfer_events WHERE token_symbol = ?", (token_symbol,))
    conn.execute("DELETE FROM token_holders WHERE token_symbol = ?", (token_symbol,))
    conn.execute(
        """
        INSERT INTO token_index_state (
            token_symbol, contract_address, decimals, start_block, last_scanned_block,
            latest_known_block, total_transfers, total_holders, last_scanned_at, last_error
        )
        VALUES (?, ?, ?, ?, ?, NULL, 0, 0, NULL, ?)
        ON CONFLICT(token_symbol) DO UPDATE SET
            contract_address = excluded.contract_address,
            decimals = excluded.decimals,
            start_block = excluded.start_block,
            last_scanned_block = excluded.last_scanned_block,
            latest_known_block = NULL,
            total_transfers = 0,
            total_holders = 0,
            last_scanned_at = NULL,
            last_error = excluded.last_error
        """,
        (token_symbol, contract_address, decimals, start_block, start_block - 1, reason),
    )


def ensure_token_index_metadata(
    conn: sqlite3.Connection,
    token_symbol: str,
    contract_address: str,
    decimals: int,
    start_block: int,
) -> dict[str, Any]:
    row = conn.execute(
        "SELECT * FROM token_index_state WHERE token_symbol = ?",
        (token_symbol,),
    ).fetchone()

    if row is None:
        reset_token_index(conn, token_symbol, contract_address, decimals, start_block)
        row = conn.execute(
            "SELECT * FROM token_index_state WHERE token_symbol = ?",
            (token_symbol,),
        ).fetchone()
        return dict(row)

    if row["contract_address"].lower() != contract_address.lower():
        reset_token_index(
            conn,
            token_symbol,
            contract_address,
            decimals,
            start_block,
            reason="Indice reconstruido por cambio de contrato.",
        )
    elif start_block < int(row["start_block"]):
        reset_token_index(
            conn,
            token_symbol,
            contract_address,
            decimals,
            start_block,
            reason="Indice reconstruido por cambio de startBlock.",
        )
    else:
        conn.execute(
            """
            UPDATE token_index_state
            SET contract_address = ?, decimals = ?, start_block = ?
            WHERE token_symbol = ?
            """,
            (contract_address, decimals, start_block, token_symbol),
        )

    row = conn.execute(
        "SELECT * FROM token_index_state WHERE token_symbol = ?",
        (token_symbol,),
    ).fetchone()
    return dict(row)


def get_index_state(token_symbol: str) -> dict[str, Any]:
    init_db()
    conn = get_db_connection()
    try:
        row = conn.execute(
            "SELECT * FROM token_index_state WHERE token_symbol = ?",
            (token_symbol,),
        ).fetchone()
        if not row:
            raise ValueError(f"No existe índice para el token {token_symbol}")
        return dict(row)
    finally:
        conn.close()


def apply_holder_delta(
    conn: sqlite3.Connection,
    token_symbol: str,
    wallet: str,
    delta: int,
    updated_at: float,
):
    if delta == 0:
        return

    row = conn.execute(
        "SELECT raw_balance FROM token_holders WHERE token_symbol = ? AND wallet = ?",
        (token_symbol, wallet),
    ).fetchone()
    current_balance = int(row["raw_balance"]) if row else 0
    new_balance = current_balance + delta

    if new_balance < 0:
        raise RuntimeError(
            f"Se detectó un balance negativo en el índice para {wallet}. "
            "Esto indica inconsistencia y requiere reconstrucción del índice."
        )

    if new_balance == 0:
        conn.execute(
            "DELETE FROM token_holders WHERE token_symbol = ? AND wallet = ?",
            (token_symbol, wallet),
        )
        return

    new_balance_text = str(new_balance)
    conn.execute(
        """
        INSERT INTO token_holders (token_symbol, wallet, raw_balance, raw_balance_length, updated_at)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(token_symbol, wallet) DO UPDATE SET
            raw_balance = excluded.raw_balance,
            raw_balance_length = excluded.raw_balance_length,
            updated_at = excluded.updated_at
        """,
        (token_symbol, wallet, new_balance_text, len(new_balance_text), updated_at),
    )


def apply_transfer_batch(
    token_symbol: str,
    contract_address: str,
    decimals: int,
    start_block: int,
    latest_known_block: int,
    last_scanned_block: int,
    logs: list,
) -> dict[str, int]:
    init_db()
    inserted_transfers = 0
    conn = get_db_connection()
    now = time.time()
    try:
        conn.execute("BEGIN IMMEDIATE")
        ensure_token_index_metadata(conn, token_symbol, contract_address, decimals, start_block)

        previous_total_transfers_row = conn.execute(
            "SELECT total_transfers FROM token_index_state WHERE token_symbol = ?",
            (token_symbol,),
        ).fetchone()
        previous_total_transfers = int(previous_total_transfers_row["total_transfers"]) if previous_total_transfers_row else 0

        for raw_log in logs:
            log = decode_transfer_log(raw_log)
            if not log:
                continue

            cursor = conn.execute(
                """
                INSERT OR IGNORE INTO token_transfer_events (
                    token_symbol, tx_hash, log_index, block_number, block_hash,
                    from_address, to_address, raw_value
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    token_symbol,
                    log["tx_hash"],
                    log["log_index"],
                    log["block_number"],
                    log["block_hash"],
                    log["from_address"],
                    log["to_address"],
                    str(log["raw_value"]),
                ),
            )

            if cursor.rowcount == 0:
                continue

            inserted_transfers += 1
            if log["from_address"] != ZERO_ADDR:
                apply_holder_delta(
                    conn,
                    token_symbol,
                    log["from_address"],
                    -int(log["raw_value"]),
                    now,
                )
            if log["to_address"] != ZERO_ADDR:
                apply_holder_delta(
                    conn,
                    token_symbol,
                    log["to_address"],
                    int(log["raw_value"]),
                    now,
                )

        total_holders_row = conn.execute(
            "SELECT COUNT(*) AS total FROM token_holders WHERE token_symbol = ?",
            (token_symbol,),
        ).fetchone()
        total_holders = int(total_holders_row["total"]) if total_holders_row else 0

        conn.execute(
            """
            UPDATE token_index_state
            SET contract_address = ?,
                decimals = ?,
                start_block = ?,
                last_scanned_block = ?,
                latest_known_block = ?,
                total_transfers = ?,
                total_holders = ?,
                last_scanned_at = ?,
                last_error = NULL
            WHERE token_symbol = ?
            """,
            (
                contract_address,
                decimals,
                start_block,
                last_scanned_block,
                latest_known_block,
                previous_total_transfers + inserted_transfers,
                total_holders,
                now,
                token_symbol,
            ),
        )
        conn.commit()
        return {
            "insertedTransfers": inserted_transfers,
            "totalHolders": total_holders,
            "totalTransfers": previous_total_transfers + inserted_transfers,
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def rewind_token_window(
    token_symbol: str,
    contract_address: str,
    decimals: int,
    start_block: int,
    latest_known_block: int,
    from_block: int,
) -> dict[str, int]:
    init_db()
    conn = get_db_connection()
    now = time.time()
    try:
        conn.execute("BEGIN IMMEDIATE")
        state = ensure_token_index_metadata(conn, token_symbol, contract_address, decimals, start_block)
        effective_from_block = max(start_block, from_block)
        current_last_scanned = int(state["last_scanned_block"])

        if effective_from_block > current_last_scanned:
            conn.commit()
            return {
                "rewindFromBlock": effective_from_block,
                "lastScannedBlock": current_last_scanned,
                "revertedTransfers": 0,
                "totalHolders": int(state["total_holders"]),
                "totalTransfers": int(state["total_transfers"]),
            }

        rows = conn.execute(
            """
            SELECT from_address, to_address, raw_value
            FROM token_transfer_events
            WHERE token_symbol = ? AND block_number >= ?
            ORDER BY block_number DESC, log_index DESC
            """,
            (token_symbol, effective_from_block),
        ).fetchall()

        reverted_transfers = 0
        for row in rows:
            raw_value = int(row["raw_value"])
            if row["from_address"] != ZERO_ADDR:
                apply_holder_delta(conn, token_symbol, row["from_address"], raw_value, now)
            if row["to_address"] != ZERO_ADDR:
                apply_holder_delta(conn, token_symbol, row["to_address"], -raw_value, now)
            reverted_transfers += 1

        conn.execute(
            "DELETE FROM token_transfer_events WHERE token_symbol = ? AND block_number >= ?",
            (token_symbol, effective_from_block),
        )

        total_transfers_row = conn.execute(
            "SELECT COUNT(*) AS total FROM token_transfer_events WHERE token_symbol = ?",
            (token_symbol,),
        ).fetchone()
        total_holders_row = conn.execute(
            "SELECT COUNT(*) AS total FROM token_holders WHERE token_symbol = ?",
            (token_symbol,),
        ).fetchone()
        total_transfers = int(total_transfers_row["total"]) if total_transfers_row else 0
        total_holders = int(total_holders_row["total"]) if total_holders_row else 0
        last_scanned_block = effective_from_block - 1

        conn.execute(
            """
            UPDATE token_index_state
            SET contract_address = ?,
                decimals = ?,
                start_block = ?,
                last_scanned_block = ?,
                latest_known_block = ?,
                total_transfers = ?,
                total_holders = ?,
                last_scanned_at = ?,
                last_error = NULL
            WHERE token_symbol = ?
            """,
            (
                contract_address,
                decimals,
                start_block,
                last_scanned_block,
                latest_known_block,
                total_transfers,
                total_holders,
                now,
                token_symbol,
            ),
        )
        conn.commit()
        return {
            "rewindFromBlock": effective_from_block,
            "lastScannedBlock": last_scanned_block,
            "revertedTransfers": reverted_transfers,
            "totalHolders": total_holders,
            "totalTransfers": total_transfers,
        }
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def update_index_latest_block(
    token_symbol: str,
    contract_address: str,
    decimals: int,
    start_block: int,
    latest_known_block: int,
):
    init_db()
    conn = get_db_connection()
    try:
        conn.execute("BEGIN IMMEDIATE")
        state = ensure_token_index_metadata(conn, token_symbol, contract_address, decimals, start_block)
        conn.execute(
            """
            UPDATE token_index_state
            SET contract_address = ?,
                decimals = ?,
                start_block = ?,
                latest_known_block = ?,
                last_error = NULL
            WHERE token_symbol = ?
            """,
            (contract_address, decimals, start_block, latest_known_block, token_symbol),
        )
        conn.commit()
        state["latest_known_block"] = latest_known_block
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def set_index_error(token_symbol: str, error_message: str):
    init_db()
    conn = get_db_connection()
    try:
        conn.execute(
            "UPDATE token_index_state SET last_error = ? WHERE token_symbol = ?",
            (error_message, token_symbol),
        )
        conn.commit()
    finally:
        conn.close()


def sync_token_index(web3: Web3, token: dict[str, Any], job_id: str | None = None) -> dict[str, Any]:
    init_db()
    token_symbol = token["symbol"]
    contract_address = Web3.to_checksum_address(token["contractAddress"])
    decimals = get_token_decimals(web3, contract_address, token.get("decimals", 18))
    start_block = max(int(token.get("startBlock", 0) or 0), 0)
    latest_block = int(web3.eth.block_number)
    target_block = max(start_block - 1, latest_block - TOKEN_SCANNER_CONFIRMATION_BLOCKS)

    conn = get_db_connection()
    try:
        conn.execute("BEGIN IMMEDIATE")
        state = ensure_token_index_metadata(conn, token_symbol, contract_address, decimals, start_block)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    original_next_block = max(start_block, int(state["last_scanned_block"]) + 1)
    rescan_from = original_next_block
    reverted_transfers = 0
    rechecked_blocks_processed = 0
    new_blocks_processed = 0
    inserted_transfers = 0

    if int(state["last_scanned_block"]) >= start_block and target_block >= start_block:
        candidate_rescan_from = max(
            start_block,
            int(state["last_scanned_block"]) - TOKEN_SCANNER_REORG_LOOKBACK_BLOCKS + 1,
        )
        if target_block >= candidate_rescan_from:
            rescan_from = candidate_rescan_from
            rewind_summary = rewind_token_window(
                token_symbol=token_symbol,
                contract_address=contract_address,
                decimals=decimals,
                start_block=start_block,
                latest_known_block=latest_block,
                from_block=rescan_from,
            )
            reverted_transfers = rewind_summary["revertedTransfers"]
            if job_id:
                update_job(
                    job_id,
                    message=(
                        f"Revalidando ventana segura desde el bloque {rewind_summary['rewindFromBlock']:,} "
                        f"antes de indexar bloques nuevos."
                    ),
                    candidateWallets=rewind_summary["totalHolders"],
                    indexedTransfers=rewind_summary["totalTransfers"],
                    recheckedBlocks=max(0, original_next_block - rescan_from),
                    revertedTransfers=reverted_transfers,
                )

    next_block = max(start_block, rescan_from)

    if target_block >= next_block:
        total_batches = ((target_block - next_block) // SCAN_BATCH_SIZE) + 1
        transfer_topic = Web3.to_hex(Web3.keccak(text="Transfer(address,address,uint256)"))
        batch_index = 0
        overlap_end = min(target_block, original_next_block - 1)

        for block_start in range(next_block, target_block + 1, SCAN_BATCH_SIZE):
            ensure_job_not_cancelled(job_id)
            block_end = min(block_start + SCAN_BATCH_SIZE - 1, target_block)
            batch_index += 1

            if job_id:
                progress = round((batch_index / max(total_batches, 1)) * 80, 2)
                update_job(
                    job_id,
                    phase="indexing",
                    progress=progress,
                    message=f"Sincronizando índice local: bloques {block_start:,}-{block_end:,}",
                )

            logs = safe_get_logs(
                web3,
                contract_address,
                transfer_topic,
                block_start,
                block_end,
                job_id,
                max_retries=6,
                min_batch=SCAN_MIN_BATCH_SIZE,
            )

            batch_summary = apply_transfer_batch(
                token_symbol=token_symbol,
                contract_address=contract_address,
                decimals=decimals,
                start_block=start_block,
                latest_known_block=latest_block,
                last_scanned_block=block_end,
                logs=logs,
            )
            inserted_transfers += batch_summary["insertedTransfers"]
            rechecked_in_batch = 0
            if overlap_end >= next_block:
                rechecked_in_batch = max(0, min(block_end, overlap_end) - block_start + 1)
            rechecked_blocks_processed += rechecked_in_batch
            new_blocks_processed += (block_end - block_start + 1) - rechecked_in_batch

            if job_id:
                update_job(
                    job_id,
                    candidateWallets=batch_summary["totalHolders"],
                    indexedTransfers=batch_summary["totalTransfers"],
                    indexedThroughBlock=block_end,
                    latestBlock=latest_block,
                    newTransfersIndexed=inserted_transfers,
                    newBlocksProcessed=new_blocks_processed,
                    recheckedBlocks=rechecked_blocks_processed,
                    revertedTransfers=reverted_transfers,
                )
    else:
        update_index_latest_block(
            token_symbol=token_symbol,
            contract_address=contract_address,
            decimals=decimals,
            start_block=start_block,
            latest_known_block=latest_block,
        )

    return {
        **get_index_state(token_symbol),
        "latestBlock": latest_block,
        "targetBlock": target_block,
        "newBlocksProcessed": new_blocks_processed,
        "recheckedBlocks": rechecked_blocks_processed,
        "revertedTransfers": reverted_transfers,
        "newTransfersIndexed": inserted_transfers,
        "cacheWarm": new_blocks_processed == 0,
    }


def decimal_token_to_raw_bound(value: Decimal, decimals: int, rounding: str) -> int:
    scale = Decimal(10) ** decimals
    return int((value * scale).to_integral_value(rounding=rounding))


def coerce_decimal(value: Decimal | str | int | float | None) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def build_raw_balance_filters(
    min_tokens: Decimal | None,
    max_tokens: Decimal | None,
    decimals: int,
) -> tuple[list[str], list[Any], int | None, int | None]:
    min_tokens = coerce_decimal(min_tokens)
    max_tokens = coerce_decimal(max_tokens)
    clauses: list[str] = []
    params: list[Any] = []
    min_raw: int | None = None
    max_raw: int | None = None

    if min_tokens is not None:
        min_raw = decimal_token_to_raw_bound(min_tokens, decimals, ROUND_CEILING)
        if min_raw > 0:
            min_text = str(min_raw)
            min_len = len(min_text)
            clauses.append("(raw_balance_length > ? OR (raw_balance_length = ? AND raw_balance >= ?))")
            params.extend([min_len, min_len, min_text])

    if max_tokens is not None:
        max_raw = decimal_token_to_raw_bound(max_tokens, decimals, ROUND_FLOOR)
        max_text = str(max(max_raw, 0))
        max_len = len(max_text)
        clauses.append("(raw_balance_length < ? OR (raw_balance_length = ? AND raw_balance <= ?))")
        params.extend([max_len, max_len, max_text])

    return clauses, params, min_raw, max_raw


def query_indexed_wallets(
    token_symbol: str,
    decimals: int,
    job_id: str | None,
    min_tokens: Decimal | None = None,
    max_tokens: Decimal | None = None,
    progress_start: float = 80.0,
    progress_end: float = 100.0,
) -> list[tuple[str, Decimal]]:
    init_db()
    clauses, params, min_raw, max_raw = build_raw_balance_filters(min_tokens, max_tokens, decimals)
    if min_raw is not None and max_raw is not None and min_raw > max_raw:
        return []

    sql_where = " WHERE token_symbol = ?"
    sql_params: list[Any] = [token_symbol]
    if clauses:
        sql_where += " AND " + " AND ".join(clauses)
        sql_params.extend(params)

    conn = get_db_connection()
    try:
        total_matches_row = conn.execute(
            f"SELECT COUNT(*) AS total FROM token_holders{sql_where}",
            sql_params,
        ).fetchone()
        total_matches = int(total_matches_row["total"]) if total_matches_row else 0
        rows = conn.execute(
            f"""
            SELECT wallet, raw_balance
            FROM token_holders
            {sql_where}
            ORDER BY raw_balance_length DESC, raw_balance DESC
            """,
            sql_params,
        ).fetchall()
    finally:
        conn.close()

    scale = Decimal(10) ** decimals
    results: list[tuple[str, Decimal]] = []
    for index, row in enumerate(rows, start=1):
        ensure_job_not_cancelled(job_id)
        if job_id and (index % 250 == 0 or index == total_matches):
            progress_span = max(progress_end - progress_start, 0)
            progress = progress_start + round((index / max(total_matches, 1)) * progress_span, 2)
            update_job(
                job_id,
                phase="filtering",
                progress=progress,
                message=f"Preparando resultados {index:,}/{total_matches:,}",
                matchesCount=index,
                scannedWallets=index,
            )
        balance = Decimal(row["raw_balance"]) / scale
        results.append((row["wallet"], balance))
    return results


def is_balance_within_range(
    balance: Decimal,
    min_tokens: Decimal | None,
    max_tokens: Decimal | None,
) -> bool:
    if min_tokens is not None and balance < min_tokens:
        return False
    if max_tokens is not None and balance > max_tokens:
        return False
    return True


def revalidate_results_with_chain(
    web3: Web3,
    token_address: str,
    decimals: int,
    indexed_results: list[tuple[str, Decimal]],
    job_id: str | None,
    min_tokens: Decimal | None = None,
    max_tokens: Decimal | None = None,
    progress_start: float = 90.0,
    progress_end: float = 99.0,
) -> tuple[list[tuple[str, Decimal]], int, int]:
    contract = web3.eth.contract(address=token_address, abi=BALANCEOF_ABI)
    verified_results: list[tuple[str, Decimal]] = []
    mismatches = 0
    removed_after_recheck = 0
    scale = Decimal(10) ** decimals
    total_wallets = len(indexed_results)
    progress_span = max(progress_end - progress_start, 0)

    for index, (wallet, indexed_balance) in enumerate(indexed_results, start=1):
        ensure_job_not_cancelled(job_id)
        if job_id and (index % 25 == 0 or index == total_wallets):
            progress = progress_start + round((index / max(total_wallets, 1)) * progress_span, 2)
            update_job(
                job_id,
                phase="validating",
                progress=progress,
                message=f"Verificando balances on-chain {index:,}/{total_wallets:,}",
                strictChecked=index,
            )

        live_raw = balance_of_with_retry(contract, wallet, job_id, max_retries=4)
        live_balance = Decimal(live_raw) / scale
        if live_balance != indexed_balance:
            mismatches += 1

        if live_raw <= 0 or not is_balance_within_range(live_balance, min_tokens, max_tokens):
            removed_after_recheck += 1
            continue

        verified_results.append((wallet, live_balance))

    verified_results.sort(key=lambda item: item[1], reverse=True)
    return verified_results, mismatches, removed_after_recheck


def generate_excel_bytes(results: list[dict[str, str]], token_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Wallets {token_name}"
    ws.append(["#", "Wallet", f"Balance ({token_name})"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for index, item in enumerate(results, start=1):
        ws.append([index, item["wallet"], Decimal(item["balance"])])
        ws.cell(row=index + 1, column=3).number_format = BALANCE_NUMBER_FORMAT
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 24
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generate_csv_bytes(results: list[dict[str, str]]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(["Wallet", "Balance", "Token"])
    for item in results:
        writer.writerow([item["wallet"], item["balance"], item["tokenSymbol"]])
    return output.getvalue().encode("utf-8-sig")


def run_scan_job(
    job_id: str,
    token_symbol: str,
    min_tokens: Decimal | None,
    max_tokens: Decimal | None,
    strict_validation: bool,
):
    started_at = time.time()
    update_job(
        job_id,
        status="running",
        phase="starting",
        progress=1,
        message="Conectando a Netsbo RPC...",
        startedAt=started_at,
    )
    token = get_token_by_symbol(token_symbol)
    try:
        if not token:
            raise ValueError(f"Token no encontrado: {token_symbol}")

        web3 = get_web3_connection()
        contract_address = Web3.to_checksum_address(token["contractAddress"])
        decimals = get_token_decimals(web3, contract_address, token.get("decimals", 18))
        start_block = max(int(token.get("startBlock", 0)), 0)

        update_job(
            job_id,
            tokenAddress=contract_address,
            decimals=decimals,
            startBlock=start_block,
        )

        index_state = sync_token_index(web3, token, job_id=job_id)
        update_job(
            job_id,
            indexedThroughBlock=index_state["last_scanned_block"],
            latestBlock=index_state["latestBlock"],
            candidateWallets=index_state["total_holders"],
            indexedTransfers=index_state["total_transfers"],
            newTransfersIndexed=index_state["newTransfersIndexed"],
            newBlocksProcessed=index_state["newBlocksProcessed"],
            recheckedBlocks=index_state["recheckedBlocks"],
            revertedTransfers=index_state["revertedTransfers"],
            message=(
                f"Índice sincronizado hasta el bloque {index_state['last_scanned_block']:,}. "
                f"Holders activos: {index_state['total_holders']:,}"
            ),
        )

        results = query_indexed_wallets(
            token_symbol=token_symbol,
            decimals=decimals,
            job_id=job_id,
            min_tokens=min_tokens,
            max_tokens=max_tokens,
            progress_start=80.0,
            progress_end=90.0 if strict_validation else 100.0,
        )

        strict_mismatches = 0
        strict_removed = 0
        strict_checked = 0
        if strict_validation and results:
            results, strict_mismatches, strict_removed = revalidate_results_with_chain(
                web3=web3,
                token_address=contract_address,
                decimals=decimals,
                indexed_results=results,
                job_id=job_id,
                min_tokens=min_tokens,
                max_tokens=max_tokens,
                progress_start=90.0,
                progress_end=99.0,
            )
            strict_checked = len(results) + strict_removed

        finished_at = time.time()
        serialized_results = serialize_results(results, token_symbol)
        update_job(
            job_id,
            status="done",
            phase="done",
            progress=100,
            message=f"Escaneo completo. {len(serialized_results):,} wallets encontradas.",
            results=serialized_results,
            total=len(serialized_results),
            pageSize=PAGE_SIZE,
            matchesCount=len(serialized_results),
            scannedWallets=len(serialized_results),
            strictMismatches=strict_mismatches,
            strictRemoved=strict_removed,
            strictChecked=strict_checked,
            finishedAt=finished_at,
            durationSeconds=round(finished_at - started_at, 2),
        )
    except JobCancelled as exc:
        finished_at = time.time()
        update_job(
            job_id,
            status="cancelled",
            phase="cancelled",
            message=str(exc),
            error=None,
            finishedAt=finished_at,
            durationSeconds=round(finished_at - started_at, 2),
        )
    except Exception as exc:
        if token:
            try:
                set_index_error(token_symbol, str(exc))
            except Exception:
                pass
        finished_at = time.time()
        update_job(
            job_id,
            status="error",
            phase="error",
            message="El escaneo falló.",
            error=f"{exc}\n\n{traceback.format_exc()}",
            finishedAt=finished_at,
            durationSeconds=round(finished_at - started_at, 2),
        )


def run_wallet_report_job(job_id: str, wallet: str, days: int):
    started_at = time.time()
    update_job(
        job_id,
        status="running",
        phase="starting",
        progress=1,
        message="Conectando a Netsbo RPC...",
        startedAt=started_at,
    )

    def progress_callback(
        phase: str,
        progress: float,
        message: str,
        metrics: dict[str, Any] | None = None,
    ):
        updates: dict[str, Any] = {
            "phase": phase,
            "progress": round(progress, 2),
            "message": message,
        }
        if metrics:
            updates.update(metrics)
        update_job(job_id, **updates)

    try:
        result = generate_wallet_report(
            wallet=wallet,
            days=days,
            batch_size=WALLET_REPORT_BATCH_SIZE,
            min_batch_size=WALLET_REPORT_MIN_BATCH_SIZE,
            progress_callback=progress_callback,
            cancel_callback=lambda: ensure_job_not_cancelled(job_id),
        )
        file_bytes = result["xlsxBytes"]
        file_path = get_wallet_report_file_path(job_id)
        temp_path = file_path.with_suffix(".xlsx.tmp")
        temp_path.write_bytes(file_bytes)
        temp_path.replace(file_path)

        finished_at = time.time()
        update_job(
            job_id,
            status="done",
            phase="done",
            progress=100,
            message=(
                f"Reporte completo: {result['transfersIn']:,} IN, {result['transfersOut']:,} OUT"
                + (f", {result['transfersSelf']:,} SELF" if result["transfersSelf"] else "")
            ),
            wallet=result["wallet"],
            days=result["days"],
            startBlock=result["startBlock"],
            endBlock=result["endBlock"],
            totalBlocks=result["totalBlocks"],
            totalTransfers=result["totalTransfers"],
            transfersIn=result["transfersIn"],
            transfersOut=result["transfersOut"],
            transfersSelf=result["transfersSelf"],
            uniqueTokens=result["uniqueTokens"],
            reportFilename=result["filename"],
            downloadUrl=build_wallet_report_download_url(job_id),
            finishedAt=finished_at,
            durationSeconds=round(finished_at - started_at, 2),
        )
        snapshot = get_job_snapshot(job_id)
        if snapshot:
            persist_wallet_report_manifest(snapshot)
    except JobCancelled as exc:
        finished_at = time.time()
        update_job(
            job_id,
            status="cancelled",
            phase="cancelled",
            progress=0,
            message=str(exc),
            error=None,
            finishedAt=finished_at,
            durationSeconds=round(finished_at - started_at, 2),
        )
        snapshot = get_job_snapshot(job_id)
        if snapshot:
            persist_wallet_report_manifest(snapshot)
    except Exception as exc:
        finished_at = time.time()
        update_job(
            job_id,
            status="error",
            phase="error",
            message="El reporte de wallet falló.",
            error=f"{exc}\n\n{traceback.format_exc()}",
            finishedAt=finished_at,
            durationSeconds=round(finished_at - started_at, 2),
        )
        snapshot = get_job_snapshot(job_id)
        if snapshot:
            persist_wallet_report_manifest(snapshot)


def get_completed_job_or_response(job_id: str):
    job = get_job_snapshot(job_id)
    if not job:
        return None, (jsonify({"error": "job_id no encontrado"}), 404)
    if job["status"] != "done":
        return None, (jsonify({"error": "El trabajo todavía no terminó"}), 409)
    return job, None


@app.route("/")
def index():
    return render_template(
        "index.html",
        strict_validation_default=STRICT_RESULT_VALIDATION_DEFAULT,
        wallet_report_max_days=WALLET_REPORT_MAX_DAYS,
    )


@app.route("/api/tokens")
def api_tokens():
    cleanup_finished_jobs()
    try:
        return jsonify(load_tokens())
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/wallet-report", methods=["POST"])
def api_wallet_report():
    cleanup_finished_jobs()
    data = request.get_json(silent=True) or {}
    wallet = str(data.get("wallet", "")).strip()
    if not wallet:
        return jsonify({"error": "wallet es obligatorio"}), 400
    if not Web3.is_address(wallet):
        return jsonify({"error": "wallet inválida"}), 400

    try:
        normalized_wallet = Web3.to_checksum_address(wallet)
    except Exception:
        return jsonify({"error": "wallet inválida"}), 400

    try:
        days = int(data.get("days", 30))
    except (TypeError, ValueError):
        return jsonify({"error": "days debe ser un entero positivo"}), 400
    if days < 1 or days > WALLET_REPORT_MAX_DAYS:
        return jsonify({"error": f"days debe estar entre 1 y {WALLET_REPORT_MAX_DAYS}"}), 400

    job_id = uuid.uuid4().hex
    with JOBS_LOCK:
        JOBS[job_id] = {
            "jobId": job_id,
            "jobType": "wallet-report",
            "status": "pending",
            "phase": "queued",
            "progress": 0,
            "message": "Trabajo en cola...",
            "error": None,
            "wallet": normalized_wallet,
            "days": days,
            "startBlock": None,
            "endBlock": None,
            "totalBlocks": None,
            "totalTransfers": 0,
            "transfersIn": 0,
            "transfersOut": 0,
            "transfersSelf": 0,
            "uniqueTokens": 0,
            "reportFilename": None,
            "downloadUrl": None,
            "createdAt": time.time(),
            "startedAt": None,
            "finishedAt": None,
            "durationSeconds": None,
            "cancelRequested": False,
        }

    thread = threading.Thread(
        target=run_wallet_report_job,
        args=(job_id, normalized_wallet, days),
        daemon=True,
    )
    thread.start()
    return jsonify({"jobId": job_id})


@app.route("/api/wallet-report/cancel", methods=["POST"])
def api_wallet_report_cancel():
    cleanup_finished_jobs()
    data = request.get_json(silent=True) or {}
    job_id = str(data.get("jobId", "")).strip()
    if not job_id:
        return jsonify({"error": "jobId es obligatorio"}), 400

    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job or job.get("jobType") != "wallet-report":
            return jsonify({"error": "job_id no encontrado"}), 404
        if job["status"] not in {"pending", "running"}:
            return jsonify({"error": "El trabajo ya no se puede cancelar"}), 409
        job["cancelRequested"] = True
        job["message"] = "Cancelación solicitada..."

    return jsonify({"jobId": job_id, "cancelRequested": True})


@app.route("/api/wallet-report/status")
def api_wallet_report_status():
    cleanup_finished_jobs()
    job_id = request.args.get("job_id", "").strip()
    if not job_id:
        return jsonify({"error": "job_id es obligatorio"}), 400

    job = get_wallet_report_job_snapshot(job_id)
    if not job:
        return jsonify({"error": "job_id no encontrado"}), 404

    return jsonify({
        "jobId": job.get("jobId"),
        "jobType": "wallet-report",
        "status": job.get("status"),
        "phase": job.get("phase"),
        "progress": job.get("progress", 0),
        "message": job.get("message"),
        "error": job.get("error"),
        "wallet": job.get("wallet"),
        "days": job.get("days"),
        "startBlock": job.get("startBlock"),
        "endBlock": job.get("endBlock"),
        "totalBlocks": job.get("totalBlocks"),
        "totalTransfers": job.get("totalTransfers", 0),
        "transfersIn": job.get("transfersIn", 0),
        "transfersOut": job.get("transfersOut", 0),
        "transfersSelf": job.get("transfersSelf", 0),
        "uniqueTokens": job.get("uniqueTokens", 0),
        "reportFilename": job.get("reportFilename"),
        "downloadUrl": job.get("downloadUrl"),
        "durationSeconds": job.get("durationSeconds"),
        "cancelRequested": bool(job.get("cancelRequested", False)),
        "finishedAt": job.get("finishedAt"),
    })


@app.route("/api/wallet-report/download")
def api_wallet_report_download():
    cleanup_finished_jobs()
    job_id = request.args.get("job_id", "").strip()
    if not job_id:
        return jsonify({"error": "job_id es obligatorio"}), 400

    job = get_wallet_report_job_snapshot(job_id)
    if not job:
        return jsonify({"error": "job_id no encontrado"}), 404
    if job.get("status") != "done":
        return jsonify({"error": "El reporte todavía no está listo"}), 409

    file_path = get_wallet_report_file_path(job_id)
    if not file_path.exists():
        return jsonify({"error": "No se encontró el archivo del reporte"}), 404

    return send_file(
        file_path,
        as_attachment=True,
        download_name=job.get("reportFilename") or f"wallet-report-{job_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/token-scan", methods=["POST"])
def api_token_scan():
    cleanup_finished_jobs()
    init_db()
    data = request.get_json(silent=True) or {}
    token_symbol = str(data.get("tokenSymbol", "")).strip()
    if not token_symbol:
        return jsonify({"error": "tokenSymbol es obligatorio"}), 400
    if not get_token_by_symbol(token_symbol):
        return jsonify({"error": f"Token no encontrado: {token_symbol}"}), 404

    try:
        min_tokens = parse_optional_decimal(data.get("minTokens"), "minTokens")
        max_tokens = parse_optional_decimal(data.get("maxTokens"), "maxTokens")
        strict_validation = parse_optional_bool(data.get("strictValidation"), "strictValidation")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    if min_tokens is not None and max_tokens is not None and min_tokens > max_tokens:
        return jsonify({"error": "minTokens no puede ser mayor que maxTokens"}), 400
    if strict_validation is None:
        strict_validation = STRICT_RESULT_VALIDATION_DEFAULT

    job_id = uuid.uuid4().hex
    with JOBS_LOCK:
        JOBS[job_id] = {
            "jobId": job_id,
            "status": "pending",
            "phase": "queued",
            "progress": 0,
            "message": "Trabajo en cola...",
            "results": [],
            "total": 0,
            "pageSize": PAGE_SIZE,
            "error": None,
            "matchesCount": 0,
            "candidateWallets": 0,
            "scannedWallets": 0,
            "indexedTransfers": 0,
            "newTransfersIndexed": 0,
            "newBlocksProcessed": 0,
            "recheckedBlocks": 0,
            "revertedTransfers": 0,
            "indexedThroughBlock": None,
            "latestBlock": None,
            "strictValidation": strict_validation,
            "strictChecked": 0,
            "strictMismatches": 0,
            "strictRemoved": 0,
            "createdAt": time.time(),
            "finishedAt": None,
            "durationSeconds": None,
            "tokenSymbol": token_symbol,
            "minTokens": str(min_tokens) if min_tokens is not None else None,
            "maxTokens": str(max_tokens) if max_tokens is not None else None,
            "cancelRequested": False,
        }

    thread = threading.Thread(
        target=run_scan_job,
        args=(job_id, token_symbol, min_tokens, max_tokens, strict_validation),
        daemon=True,
    )
    thread.start()
    return jsonify({"jobId": job_id})


@app.route("/api/token-scan/cancel", methods=["POST"])
def api_token_scan_cancel():
    cleanup_finished_jobs()
    data = request.get_json(silent=True) or {}
    job_id = str(data.get("jobId", "")).strip()
    if not job_id:
        return jsonify({"error": "jobId es obligatorio"}), 400

    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return jsonify({"error": "job_id no encontrado"}), 404
        if job["status"] not in {"pending", "running"}:
            return jsonify({"error": "El trabajo ya no se puede cancelar"}), 409
        job["cancelRequested"] = True
        job["message"] = "Cancelación solicitada..."

    return jsonify({"jobId": job_id, "cancelRequested": True})


@app.route("/api/token-scan/status")
def api_token_scan_status():
    cleanup_finished_jobs()
    job_id = request.args.get("job_id", "").strip()
    job = get_job_snapshot(job_id)
    if not job:
        return jsonify({"error": "job_id no encontrado"}), 404
    return jsonify({
        "jobId": job["jobId"],
        "status": job["status"],
        "phase": job.get("phase"),
        "progress": job.get("progress", 0),
        "message": job.get("message"),
        "error": job.get("error"),
        "matchesCount": job.get("matchesCount", 0),
        "candidateWallets": job.get("candidateWallets", 0),
        "scannedWallets": job.get("scannedWallets", 0),
        "indexedTransfers": job.get("indexedTransfers", 0),
        "newTransfersIndexed": job.get("newTransfersIndexed", 0),
        "newBlocksProcessed": job.get("newBlocksProcessed", 0),
        "recheckedBlocks": job.get("recheckedBlocks", 0),
        "revertedTransfers": job.get("revertedTransfers", 0),
        "indexedThroughBlock": job.get("indexedThroughBlock"),
        "latestBlock": job.get("latestBlock"),
        "strictValidation": job.get("strictValidation", False),
        "strictChecked": job.get("strictChecked", 0),
        "strictMismatches": job.get("strictMismatches", 0),
        "strictRemoved": job.get("strictRemoved", 0),
        "durationSeconds": job.get("durationSeconds"),
        "cancelRequested": job.get("cancelRequested", False),
        "finishedAt": job.get("finishedAt"),
    })


@app.route("/api/token-scan/result")
def api_token_scan_result():
    cleanup_finished_jobs()
    job_id = request.args.get("job_id", "").strip()
    try:
        page = parse_page_param(request.args.get("page", 1))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    job, error_response = get_completed_job_or_response(job_id)
    if error_response:
        return error_response

    all_results = job.get("results", [])
    total = job.get("total", len(all_results))
    page_size = job.get("pageSize", PAGE_SIZE)
    total_pages = max(1, (total + page_size - 1) // page_size)
    start = (page - 1) * page_size
    end = start + page_size

    return jsonify({
        "results": all_results[start:end],
        "total": total,
        "page": page,
        "pageSize": page_size,
        "totalPages": total_pages,
        "durationSeconds": job.get("durationSeconds"),
        "candidateWallets": job.get("candidateWallets", 0),
        "scannedWallets": job.get("scannedWallets", 0),
        "indexedTransfers": job.get("indexedTransfers", 0),
        "newTransfersIndexed": job.get("newTransfersIndexed", 0),
        "newBlocksProcessed": job.get("newBlocksProcessed", 0),
        "recheckedBlocks": job.get("recheckedBlocks", 0),
        "revertedTransfers": job.get("revertedTransfers", 0),
        "indexedThroughBlock": job.get("indexedThroughBlock"),
        "latestBlock": job.get("latestBlock"),
        "strictValidation": job.get("strictValidation", False),
        "strictChecked": job.get("strictChecked", 0),
        "strictMismatches": job.get("strictMismatches", 0),
        "strictRemoved": job.get("strictRemoved", 0),
        "cached": job.get("newBlocksProcessed", 0) == 0,
    })


@app.route("/api/token-scan/export")
def api_token_scan_export():
    cleanup_finished_jobs()
    job_id = request.args.get("job_id", "").strip()
    job, error_response = get_completed_job_or_response(job_id)
    if error_response:
        return error_response

    results = job.get("results", [])
    if not results:
        return jsonify({"error": "No hay archivo para exportar"}), 404

    token_symbol = job.get("tokenSymbol", "token")
    file_bytes = generate_excel_bytes(results, token_symbol)
    return send_file(
        io.BytesIO(file_bytes),
        as_attachment=True,
        download_name=f"token-scan-{token_symbol}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/token-scan/export.csv")
def api_token_scan_export_csv():
    cleanup_finished_jobs()
    job_id = request.args.get("job_id", "").strip()
    job, error_response = get_completed_job_or_response(job_id)
    if error_response:
        return error_response

    results = job.get("results", [])
    if not results:
        return jsonify({"error": "No hay archivo para exportar"}), 404

    token_symbol = job.get("tokenSymbol", "token")
    file_bytes = generate_csv_bytes(results)
    return send_file(
        io.BytesIO(file_bytes),
        as_attachment=True,
        download_name=f"token-scan-{token_symbol}.csv",
        mimetype="text/csv",
    )


init_db()


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=DEBUG_MODE)
